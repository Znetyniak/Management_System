# app/controllers/store_controller.py

from flask import (
    render_template, request, jsonify,
    redirect, url_for, flash, session, send_file
)
from functools import wraps

from app.controllers import store_bp

from app.services.vehicle_service import VehicleService
from app.services.driver_service import DriverService
from app.services.trip_service import TripService
from app.services.expense_service import ExpenseService
from app.services.maintenance_service import MaintenanceService

from app.reports.usage_report import UsageReport

from datetime import datetime
from openpyxl import Workbook
import tempfile


# ---------------------------------------------------------
# ACCESS CONTROL — ADMIN OR USER
# ---------------------------------------------------------
def store_access_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            flash("You must log in first.", "warning")
            return redirect(url_for("auth.login"))

        if session.get("role") not in ("user", "admin"):
            flash("Access denied.", "danger")
            return redirect(url_for("auth.login"))

        return func(*args, **kwargs)
    return wrapper


# ---------------------------------------------------------
# Helper: safe parse date
# ---------------------------------------------------------
def _safe_parse_date(x):
    if not x:
        return None
    if isinstance(x, datetime):
        return x
    if isinstance(x, str):
        for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(x, fmt)
            except:
                pass
        try:
            return datetime.fromisoformat(x)
        except:
            return None
    return None


# ---------------------------------------------------------
# STORE DASHBOARD
# ---------------------------------------------------------
@store_bp.route('/dashboard')
@store_access_required
def dashboard():
    vehicles_raw = VehicleService.get_all()
    all_vehicles = vehicles_raw[0] if isinstance(vehicles_raw, tuple) else vehicles_raw
    total_vehicles = len(all_vehicles)

    active_trips = TripService.count_active()
    upcoming_maint = MaintenanceService.count_upcoming()
    month_expenses = ExpenseService.total_month()

    return render_template(
        "Store/dashboard.html",
        total_vehicles=total_vehicles,
        active_trips=active_trips,
        upcoming_maint=upcoming_maint,
        month_expenses=month_expenses
    )


# ---------------------------------------------------------
# VEHICLES LIST
# ---------------------------------------------------------
@store_bp.route('/vehicles')
@store_access_required
def vehicles():
    q = request.args.get("q", "").strip().lower()
    page = int(request.args.get("page", 1))
    per_page = 10

    vehicles_raw = VehicleService.get_all()
    all_vehicles = vehicles_raw[0] if isinstance(vehicles_raw, tuple) else vehicles_raw

    if q:
        all_vehicles = [
            v for v in all_vehicles
            if q in v.brand.lower()
            or q in v.model.lower()
            or q in v.vin.lower()
        ]

    total = len(all_vehicles)
    start = (page - 1) * per_page
    end = start + per_page

    vehicles_page = all_vehicles[start:end]
    pages = (total + per_page - 1) // per_page

    return render_template(
        "Store/vehicles.html",
        vehicles=vehicles_page,
        page=page,
        pages=pages,
        has_next=end < total,
        q=q
    )


# ---------------------------------------------------------
# VEHICLE DETAIL
# ---------------------------------------------------------
@store_bp.route('/vehicle/<int:vid>')
@store_access_required
def vehicle_detail(vid):
    vehicle = VehicleService.get_by_id(vid)
    if not vehicle:
        flash("Vehicle not found.", "danger")
        return redirect(url_for("store.vehicles"))
    return render_template("Store/vehicle_detail.html", vehicle=vehicle)


# ---------------------------------------------------------
# DRIVERS LIST
# ---------------------------------------------------------
@store_bp.route('/drivers')
@store_access_required
def drivers():
    drivers_raw = DriverService.get_all()
    drivers = drivers_raw[0] if isinstance(drivers_raw, tuple) else drivers_raw
    return render_template("Store/drivers.html", drivers=drivers)


# ---------------------------------------------------------
# DRIVER DETAIL
# ---------------------------------------------------------
@store_bp.route('/driver/<int:did>')
@store_access_required
def driver_detail(did):
    driver = DriverService.get_by_id(did)
    if not driver:
        flash("Driver not found.", "danger")
        return redirect(url_for("store.drivers"))
    return render_template("Store/driver_detail.html", driver=driver)


# ---------------------------------------------------------
# TRIPS LIST
# ---------------------------------------------------------
@store_bp.route('/trips')
@store_access_required
def trips():
    trips_raw = TripService.get_all_trips()
    trips = trips_raw[0] if isinstance(trips_raw, tuple) else trips_raw
    return render_template("Store/trips.html", trips=trips)


# ---------------------------------------------------------
# TRIP DETAIL
# ---------------------------------------------------------
@store_bp.route('/trip/<int:tid>')
@store_access_required
def trip_detail(tid):
    trip = TripService.get_by_id(tid)
    if not trip:
        flash("Trip not found.", "danger")
        return redirect(url_for("store.trips"))
    return render_template("Store/trip_detail.html", trip=trip)


# ---------------------------------------------------------
# EXPENSES LIST
# ---------------------------------------------------------
@store_bp.route('/expenses')
@store_access_required
def expenses():
    raw = ExpenseService.get_all()
    expenses = raw[0] if isinstance(raw, tuple) else raw
    return render_template("Store/expenses.html", expenses=expenses)


# ---------------------------------------------------------
# MAINTENANCE LIST
# ---------------------------------------------------------
@store_bp.route('/maintenance')
@store_access_required
def maintenance():
    raw = MaintenanceService.get_all()
    maint = raw[0] if isinstance(raw, tuple) else raw
    return render_template("Store/maintenance.html", maintenances=maint)


# ---------------------------------------------------------
# USAGE REPORT
# ---------------------------------------------------------
@store_bp.route('/reports/usage')
@store_access_required
def report_usage():

    start_raw = request.args.get("from")
    end_raw = request.args.get("to")
    vehicle_id_raw = request.args.get("vehicle_id")

    vehicles_raw = VehicleService.get_all()
    vehicles = vehicles_raw[0] if isinstance(vehicles_raw, tuple) else vehicles_raw

    report = UsageReport.vehicle_utilization(
        start_date=start_raw,
        end_date=end_raw,
        vehicle_id=vehicle_id_raw
    )

    return render_template(
        "Store/reports_usage.html",
        report=report,
        vehicles=vehicles,
        start_date=start_raw,
        end_date=end_raw,
        vehicle_id=vehicle_id_raw
    )


# ---------------------------------------------------------
# USAGE REPORT EXPORT
# ---------------------------------------------------------
@store_bp.route('/reports/usage/export')
@store_access_required
def report_usage_export():

    start_raw = request.args.get("from")
    end_raw = request.args.get("to")
    vehicle_id_raw = request.args.get("vehicle_id")

    report = UsageReport.vehicle_utilization(
        start_date=start_raw,
        end_date=end_raw,
        vehicle_id=vehicle_id_raw
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Usage Report"

    ws.append(["Vehicle ID", "Brand", "Model", "VIN", "Trips", "Total Distance"])

    for r in report:
        ws.append([
            r["vehicle_id"],
            r["brand"],
            r["model"],
            r["vin"],
            r["trip_count"],
            r["total_distance"]
        ])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="usage_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------------------------------------------------
# EXPENSES REPORT
# ---------------------------------------------------------
@store_bp.route('/reports/expenses')
@store_access_required
def report_expenses():

    start_raw = request.args.get("from")
    end_raw = request.args.get("to")
    vehicle_id_raw = request.args.get("vehicle_id")
    expense_type_raw = request.args.get("type")

    start = _safe_parse_date(start_raw)
    end = _safe_parse_date(end_raw)

    raw = ExpenseService.get_all()
    expenses = raw[0] if isinstance(raw, tuple) else raw

    filtered = []
    total_amount = 0
    per_vehicle = {}
    per_type = {}

    try:
        vehicle_id = int(vehicle_id_raw) if vehicle_id_raw else None
    except:
        vehicle_id = None

    for e in expenses:
        e_date = _safe_parse_date(e.date)

        if vehicle_id and e.vehicle_id != vehicle_id:
            continue

        if expense_type_raw and e.expense_type.lower() != expense_type_raw.lower():
            continue

        if start and e_date and e_date < start:
            continue
        if end and e_date and e_date > end:
            continue

        filtered.append(e)
        amount = float(e.amount or 0)
        total_amount += amount

        per_vehicle[e.vehicle_id] = per_vehicle.get(e.vehicle_id, 0) + amount
        per_type[e.expense_type] = per_type.get(e.expense_type, 0) + amount

    vehicles_raw = VehicleService.get_all()
    vehicles = vehicles_raw[0] if isinstance(vehicles_raw, tuple) else vehicles_raw

    return render_template(
        "Store/reports_expenses.html",
        expenses=filtered,
        total_amount=total_amount,
        per_vehicle=per_vehicle,
        per_type=per_type,
        vehicles=vehicles,
        start_date=start_raw,
        end_date=end_raw,
        selected_vehicle=vehicle_id_raw,
        selected_type=expense_type_raw
    )


# ---------------------------------------------------------
# EXPENSES EXPORT
# ---------------------------------------------------------
@store_bp.route('/reports/expenses/export')
@store_access_required
def report_expenses_export():

    start_raw = request.args.get("from")
    end_raw = request.args.get("to")
    vehicle_id_raw = request.args.get("vehicle_id")
    expense_type_raw = request.args.get("type")

    start = _safe_parse_date(start_raw)
    end = _safe_parse_date(end_raw)

    raw = ExpenseService.get_all()
    expenses = raw[0] if isinstance(raw, tuple) else raw

    filtered = []
    try:
        vehicle_id = int(vehicle_id_raw) if vehicle_id_raw else None
    except:
        vehicle_id = None

    for e in expenses:
        e_date = _safe_parse_date(e.date)

        if vehicle_id and e.vehicle_id != vehicle_id:
            continue
        if expense_type_raw and e.expense_type.lower() != expense_type_raw.lower():
            continue
        if start and e_date and e_date < start:
            continue
        if end and e_date and e_date > end:
            continue

        filtered.append(e)

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses Report"

    ws.append(["ID", "Vehicle ID", "Type", "Amount", "Date", "Description"])

    for e in filtered:
        e_date = _safe_parse_date(e.date)
        ws.append([
            e.id,
            e.vehicle_id,
            e.expense_type,
            e.amount,
            e_date.strftime("%Y-%m-%d") if e_date else "",
            e.description or ""
        ])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="expenses_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------------------------------------------------
# MAINTENANCE REPORT
# ---------------------------------------------------------
@store_bp.route('/reports/maintenance')
@store_access_required
def report_maintenance():

    start_raw = request.args.get("from")
    end_raw = request.args.get("to")
    vehicle_id_raw = request.args.get("vehicle_id")

    start = _safe_parse_date(start_raw)
    end = _safe_parse_date(end_raw)

    raw = MaintenanceService.get_all()
    maint = raw[0] if isinstance(raw, tuple) else raw

    try:
        vehicle_id = int(vehicle_id_raw) if vehicle_id_raw else None
    except:
        vehicle_id = None

    filtered = []
    for m in maint:
        m_date = _safe_parse_date(m.date)

        if vehicle_id and m.vehicle_id != vehicle_id:
            continue
        if start and m_date and m_date < start:
            continue
        if end and m_date and m_date > end:
            continue

        filtered.append(m)

    # ---------- BUILD STATISTICS ----------
    per_vehicle = {}
    per_type = {}

    for m in filtered:
        per_vehicle[m.vehicle_id] = per_vehicle.get(m.vehicle_id, 0) + 1
        per_type[m.maintenance_type] = per_type.get(m.maintenance_type, 0) + 1

    report = {
        "records": filtered,
        "count": len(filtered),
        "per_vehicle": per_vehicle,
        "per_type": per_type
    }

    return render_template(
        "Store/reports_maintenance.html",
        report=report,
        start_date=start_raw,
        end_date=end_raw,
        vehicle_id=vehicle_id_raw
    )


# ---------------------------------------------------------
# MAINTENANCE REPORT EXPORT
# ---------------------------------------------------------
@store_bp.route('/reports/maintenance/export')
@store_access_required
def report_maintenance_export():

    start_raw = request.args.get("from")
    end_raw = request.args.get("to")
    vehicle_id_raw = request.args.get("vehicle_id")

    start = _safe_parse_date(start_raw)
    end = _safe_parse_date(end_raw)

    raw = MaintenanceService.get_all()
    maint = raw[0] if isinstance(raw, tuple) else raw

    try:
        vehicle_id = int(vehicle_id_raw) if vehicle_id_raw else None
    except:
        vehicle_id = None

    filtered = []
    for m in maint:
        m_date = _safe_parse_date(m.date)

        if vehicle_id and m.vehicle_id != vehicle_id:
            continue
        if start and m_date and m_date < start:
            continue
        if end and m_date and m_date > end:
            continue

        filtered.append(m)

    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenance Report"

    ws.append(["ID", "Vehicle ID", "Type", "Date", "Next Date"])

    for m in filtered:
        m_date = _safe_parse_date(m.date)
        next_date = _safe_parse_date(m.next_date)

        ws.append([
            m.id,
            m.vehicle_id,
            m.maintenance_type,
            m_date.strftime("%Y-%m-%d") if m_date else "",
            next_date.strftime("%Y-%m-%d") if next_date else ""
        ])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="maintenance_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
